from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = 'jobs.xlsx'

# Ensure the Excel file exists with headers
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Company Name', 'Salary', 'Date Applied', 'Decision Status'])
    wb.save(EXCEL_FILE)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        company_name = request.form['company_name']
        salary = request.form['salary']
        date_applied = request.form['date_applied']
        decision_status = request.form['decision_status']

        # Set salary to "Not Specified" if empty
        if not salary:
            salary = "Not Specified"
        else:
            # Add "ZAR" prefix to salary
            salary = f"ZAR {salary}"

        # Load the Excel file
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        # Append the new job application
        ws.append([company_name, salary, date_applied, decision_status])

        # Save the Excel file
        wb.save(EXCEL_FILE)

        return redirect(url_for('index'))

    # Read the Excel file to display its contents
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        jobs.append(row)

    return render_template('index.html', jobs=jobs)

@app.route('/delete/<int:row_index>', methods=['POST'])
def delete_job(row_index):
    # Load the Excel file
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Delete the specified row (row_index is 0-based for the table, but Excel rows are 1-based)
    ws.delete_rows(row_index + 2)  # +2 to account for header row and 0-based index

    # Save the Excel file
    wb.save(EXCEL_FILE)

    return redirect(url_for('index'))

@app.route('/update_status/<int:row_index>', methods=['POST'])
def update_status(row_index):
    new_status = request.form['new_status']

    # Load the Excel file
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Update the decision status for the specified row
    ws.cell(row=row_index + 2, column=4).value = new_status  # Column 4 is Decision Status

    # Save the Excel file
    wb.save(EXCEL_FILE)

    return jsonify(success=True)

@app.route('/clear', methods=['POST'])
def clear_all():
    # Load the Excel file
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Clear all rows except the header
    ws.delete_rows(2, ws.max_row)  # Delete from row 2 to the last row

    # Save the Excel file
    wb.save(EXCEL_FILE)

    return redirect(url_for('index'))

@app.route('/download')
def download_file():
    # Send the Excel file for download
    return send_file(EXCEL_FILE, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)